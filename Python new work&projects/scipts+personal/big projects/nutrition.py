# to do
#  gui modules for user sort requests, and amount requests, search bar, can add items and show added items, along with current total nutrition
# checks dv values you are low on 
# can search and add food to your total nutrition for day then reccomends by lowest dvs, suggestions while searching? like type lemon shows lemonaide lemon pie etc
# can add multiple of an item without searching everytime, maybe add how many on search + a '+' '-' sign next to items to remove add quickly
# need total dv values to compare current percent of dv
# mention surpluss?
# source of data and 2015, nutition profiles as approximates region an locally grown food may change values
# Async?, use text formatter?


import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

root = tk.Tk()
root.title("Nutrition")
root.geometry("1800x800")
root.configure(bg="#f4fdfe")
root.resizable(True,True)



#opening/loading openpyxl book and sheet
wb= load_workbook('ABBREV.xlsx', read_only=True)#read only greatly improves peformance, even though its still kinda slow
ws = wb.active

#treeview frame
frame1 = tk.LabelFrame(root, text="Nutritional Data")
frame1.place(height=500, width=1800)

#frame for options
frame2 = tk.LabelFrame(root)
frame2.place(x=0, y=501)

#treeview widget
tv1= ttk.Treeview(frame1)
tv1.place(relheight=1, relwidth=1)


treescrolly=tk.Scrollbar(frame1,orient="vertical",command=tv1.yview)
treescrollx=tk.Scrollbar(frame1,orient="horizontal",command=tv1.xview)
tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set)
treescrollx.pack(side="bottom", fill="x")
treescrolly.pack(side="right", fill="y")
# a bit of a jumble but writing the data out like this, foodgroup: range of addresses in the data makes it so i dont have to do a ton of if statements or assignements for any possible click, rather inherintely the click already means what it should, very happy to not do 25 if-else checks
foodgroups = {"All": (2,8791), "Dairy and Egg Products": (2,251), "Spices and Herbs": (252,315),"Baby Foods": (316,634),"Fats and Oils" : (635,823),"Poultry Product": (824,1205),"Soups, Sauces:, and Gravies":(1206,1620),"Sausages and Luncheon Meats":(1621,1775),"Breakfast Cereals":(1776,2120),"Fruits and Fruit Juices":(2121,2479),"Pork Products":(2480,2815),"Vegetables and Vegetable Products":(2816,3606),"Nut and Seed Products":(3607,3743),"Beef Products":(3744,4123),"Beverages":(4124,4483),"Finfish and Shellfish Products":(4484,4742),"Legumes and Legume Products":(4743,5099),"Lamb, Veal, and Game Products":(5100,5563),"Baked Products":(5564,6049),"Sweet":(6050,6473),"Cereal Grains and Pasta":(6474,6653),"Fast Foods":(6654,7014),"Meals, Entrees, and Side Dishes":(7015,7112),"Snacks":(7700,7759),"American Indian/Alaska Native Foods":(8289,8453),"Restaurant Foods":(8454,8562)}


def option(*args): 
    bottom, top = 0, 0
    x = foodgroups[clicked.get()]
    bottom, top = x
    r_set=ws.iter_rows(min_row=bottom, max_row=top, min_col=2,max_col=53, values_only=True)
    r_set=[r for r in r_set]
    return r_set
    



#scrollable dropdown box for catagory selections
clicked= tk.StringVar()
clicked.set("All")
dropdown = tk.OptionMenu(frame2, clicked, *foodgroups, command=option)
dropdown.pack(side= tk.LEFT, fill =tk.BOTH)


def treeview_sort_column(tv, col, reverse):
    l = [[tv1.set(k, col), k] for k in tv1.get_children('')]
    for item in range(len(l)): #some cells in the data are just blank, no zeros, so python returns me "None" this makes it so Nones are 0s for sorting
        if l[item][0]== 'None':
            l[item][0] = 0
            
    l.sort(reverse=reverse, key=lambda tup: float(tup[0]))

    # move items
    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    # reverse the sort next click
    tv1.heading(col, text=col, command=lambda _col=col: \
                 treeview_sort_column(tv1, _col, not reverse))



row1=ws.iter_rows(min_row=1, max_row=1, min_col = 2, max_col=53, values_only=True)#using the built in iter_rows greatly improved peformance
row1=[r for r in row1]#header list
def populate():
    #assigning headers, filling treeview sheet
    tv1["column"] = row1[0]
    tv1["show"] = "headings"
    #tv1.column("Shrt_Desc", width=450,minwidth=450, stretch=True)
    for column in tv1["columns"]:
        tv1.heading(column, text=column, command=lambda _col=column:treeview_sort_column(tv1, _col, False))

    for dt in option():
        tv1.insert("","end", values=dt)

def clear_data():
    tv1.delete(*tv1.get_children())

B1 = tk.Button(frame2, text="Populate", command=lambda: [clear_data(), populate()])
B1.pack()






root.mainloop()
wb.close()